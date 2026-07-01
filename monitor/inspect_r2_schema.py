#!/usr/bin/env python3
"""Validate Motorgy Excel files uploaded to Cloudflare R2 against websites-config.yml schema."""

from __future__ import annotations

import argparse
import fnmatch
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import boto3
import numpy as np
import pandas as pd
import yaml
from botocore.config import Config
from botocore.exceptions import ClientError
from openpyxl import load_workbook

_MONITOR_DIR = os.path.dirname(os.path.abspath(__file__))
if _MONITOR_DIR not in sys.path:
    sys.path.insert(0, _MONITOR_DIR)

from ads_counter import count_scraper_ads
from github_workflows import build_scraper_run_meta, load_site_run_meta
from r2_file_counter import count_scraper_r2_files, count_site_r2_files
from request_metrics import (
    aggregate_site_request_metrics,
    build_run_error_summary,
    count_scraper_request_metrics,
)

LOCAL_CONFIG_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "websites-config.yml")
MONITOR_STATS_KEY = "monitor/monitor_stats.yml"
CONFIG_KEY = "monitor/websites-config.yml"
DEFAULT_CONFIG_BASE = "motorgy"
DEFAULT_DYNAMIC_COLUMN_PATTERNS = ["inspection_*__*"]


def json_safe(value: Any) -> Any:
    """Convert numpy/pandas scalars to native Python types for JSON serialization."""
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(v) for v in value]
    if isinstance(value, np.bool_):
        return bool(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def get_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def build_r2_client():
    return boto3.client(
        "s3",
        endpoint_url=get_env("CF_R2_ENDPOINT_URL"),
        aws_access_key_id=get_env("CF_R2_ACCESS_KEY_ID"),
        aws_secret_access_key=get_env("CF_R2_SECRET_ACCESS_KEY"),
        region_name="us-east-1",
        config=Config(signature_version="s3v4", s3={"addressing_style": "path"}),
    )


def strip_bucket_placeholder(r2_path: str) -> str:
    return re.sub(r"^\{bucket\}/?", "", r2_path).rstrip("/")


def date_partition_prefix(base_path: str, target_date: datetime) -> str:
    return (
        f"{base_path}/year={target_date.strftime('%Y')}"
        f"/month={target_date.strftime('%m')}"
        f"/day={target_date.strftime('%d')}"
    )


def list_xlsx_objects(client, bucket: str, prefix: str) -> List[dict]:
    objects: List[dict] = []
    paginator = client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            if key.lower().endswith(".xlsx"):
                objects.append(obj)
    return objects


def download_object(client, bucket: str, key: str) -> bytes:
    response = client.get_object(Bucket=bucket, Key=key)
    return response["Body"].read()


def load_config_from_r2(client, bucket: str, base_path: str = DEFAULT_CONFIG_BASE) -> dict:
    """Load websites-config.yml from R2 at {base_path}/monitor/websites-config.yml."""
    key = f"{base_path.rstrip('/')}/{CONFIG_KEY}"
    try:
        body = download_object(client, bucket, key)
        print(f"Loaded config from r2://{bucket}/{key}")
        return yaml.safe_load(body) or {}
    except ClientError as exc:
        if exc.response["Error"]["Code"] not in ("NoSuchKey", "404"):
            raise
        if os.path.isfile(LOCAL_CONFIG_PATH):
            print(
                f"WARNING: Config not found at r2://{bucket}/{key}, "
                f"using local fallback {LOCAL_CONFIG_PATH}",
                file=sys.stderr,
            )
            with open(LOCAL_CONFIG_PATH, encoding="utf-8") as fh:
                return yaml.safe_load(fh) or {}
        raise RuntimeError(
            f"websites-config.yml not found at r2://{bucket}/{key} "
            f"and no local fallback at {LOCAL_CONFIG_PATH}"
        ) from exc


def inspect_workbook(content: bytes) -> Dict[str, Any]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheets: Dict[str, Any] = {}
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            headers = [str(cell).strip() if cell is not None else "" for cell in next(rows, ())]
            data_rows = max(ws.max_row - 1, 0) if ws.max_row else 0
            sheets[sheet_name] = {"columns": headers, "row_count": data_rows}
    finally:
        wb.close()
    return {"sheets": sheets}


def classify_columns(
    columns: List[str],
    sheet_schema: dict,
    known_union: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Split columns into required, dynamic, and unexpected groups."""
    required = sheet_schema.get("required_columns", [])
    required_set = set(required)
    patterns = sheet_schema.get("dynamic_column_patterns", DEFAULT_DYNAMIC_COLUMN_PATTERNS)
    actual = list(columns)

    dynamic = [
        col
        for col in actual
        if col not in required_set
        and any(fnmatch.fnmatch(col, pattern) for pattern in patterns)
    ]
    dynamic_set = set(dynamic)
    unknown = [col for col in actual if col not in required_set and col not in dynamic_set]
    missing_required = [col for col in required if col not in actual]

    union_set = set(known_union or [])
    required_in_union = [col for col in union_set if col in required_set]
    new_columns = [col for col in actual if union_set and col not in union_set]
    dropped_from_union = (
        [col for col in required_in_union if col not in actual] if required_in_union else []
    )

    return {
        "total_columns": len(actual),
        "required_count": len(required_set & set(actual)),
        "dynamic_count": len(dynamic),
        "columns": sorted(actual),
        "required_columns": sorted(required_set & set(actual)),
        "missing_required": sorted(missing_required),
        "dynamic_columns": sorted(dynamic),
        "unknown_columns": sorted(unknown),
        "new_columns": sorted(new_columns),
        "dropped_from_union": sorted(dropped_from_union),
        "dynamic_column_patterns": patterns,
    }


def validate_column_schema(
    actual_cols: List[str],
    sheet_schema: dict,
    known_union: List[str],
) -> Tuple[List[dict], Dict[str, Any]]:
    """Validate and describe the column schema for one sheet."""
    checks: List[dict] = []
    schema = classify_columns(actual_cols, sheet_schema, known_union)
    patterns = schema["dynamic_column_patterns"]

    missing = schema["missing_required"]
    checks.append(
        {
            "check": "required_columns",
            "passed": len(missing) == 0,
            "detail": (
                f"Missing required columns: {missing}"
                if missing
                else f"All {len(sheet_schema.get('required_columns', []))} required columns present"
            ),
            "severity": "critical",
        }
    )

    unknown = schema["unknown_columns"]
    checks.append(
        {
            "check": "column_schema",
            "passed": len(unknown) == 0,
            "detail": (
                f"Unexpected columns outside schema: {unknown}"
                if unknown
                else (
                    f"Column schema OK — {schema['required_count']} required, "
                    f"{schema['dynamic_count']} dynamic ({', '.join(patterns)})"
                )
            ),
            "severity": "high",
        }
    )

    if known_union:
        dropped = schema["dropped_from_union"]
        checks.append(
            {
                "check": "column_union",
                "passed": len(dropped) == 0,
                "detail": (
                    f"Required columns missing vs R2 history: {dropped}"
                    if dropped
                    else f"All {len(sheet_schema.get('required_columns', []))} required columns match history"
                ),
                "severity": "medium",
            }
        )

    return checks, schema


def expand_file_pattern(pattern: str, target_date: datetime) -> str:
    """Replace template placeholders before fnmatch (matches CF/scrape_motorgy.py naming)."""
    return (
        pattern.replace("{date}", target_date.strftime("%Y%m%d"))
        .replace("{date_dash}", target_date.strftime("%Y-%m-%d"))
    )


def matches_file_pattern(filename: str, pattern: str, target_date: Optional[datetime] = None) -> bool:
    resolved = expand_file_pattern(pattern, target_date) if target_date else pattern
    return fnmatch.fnmatch(filename, resolved)


def normalize_part_key(filename: str) -> str:
    match = re.search(r"_part-([^./]+)", filename, re.IGNORECASE)
    if match:
        return f"part-{match.group(1)}"
    return re.sub(r"_\d{8}", "_DATE", filename)


def resolve_row_range(
    part_key: str,
    sheet_name: str,
    scraper_name: str,
    stats: dict,
    peer_row_counts: List[int],
) -> Tuple[int, Optional[int], str]:
    """Derive acceptable row range from R2 historical stats or same-run peers."""
    part_stats = (
        stats.get(scraper_name, {})
        .get("parts", {})
        .get(part_key, {})
        .get(sheet_name, {})
    )

    if part_stats.get("min_rows") is not None and part_stats.get("max_rows") is not None:
        min_rows = max(1, int(part_stats["min_rows"] * 0.8))
        max_rows = max(min_rows, int(part_stats["max_rows"] * 1.2))
        return min_rows, max_rows, "r2_historical_stats"

    non_zero_peers = [count for count in peer_row_counts if count > 0]
    if len(non_zero_peers) >= 2:
        peer_min = min(non_zero_peers)
        peer_max = max(non_zero_peers)
        min_rows = max(1, int(peer_min * 0.7))
        max_rows = max(min_rows, int(peer_max * 1.3))
        return min_rows, max_rows, "same_run_peers"

    if len(non_zero_peers) == 1:
        return 1, None, "single_file_baseline"

    return 1, None, "no_baseline"


def resolve_min_size_kb(
    part_key: str,
    sheet_name: str,
    scraper_name: str,
    stats: dict,
    peer_sizes_kb: List[float],
) -> Tuple[float, str]:
    part_stats = (
        stats.get(scraper_name, {})
        .get("parts", {})
        .get(part_key, {})
        .get(sheet_name, {})
    )

    if part_stats.get("min_size_kb") is not None:
        return max(1.0, float(part_stats["min_size_kb"]) * 0.8), "r2_historical_stats"

    non_zero_peers = [size for size in peer_sizes_kb if size > 0]
    if non_zero_peers:
        return max(1.0, min(non_zero_peers) * 0.5), "same_run_peers"

    return 1.0, "no_baseline"


def validate_file(
    filename: str,
    file_size_bytes: int,
    inspection: Dict[str, Any],
    schema_entry: dict,
    scraper_name: str,
    stats: dict,
    peer_row_counts: List[int],
    peer_sizes_kb: List[float],
    target_date: datetime,
) -> Tuple[List[dict], bool, Dict[str, Any]]:
    checks: List[dict] = []
    all_passed = True

    def add_check(name: str, passed: bool, detail: str, severity: str = "critical") -> None:
        nonlocal all_passed
        passed = bool(passed)
        if not passed:
            all_passed = False
        checks.append({"check": name, "passed": passed, "detail": detail, "severity": severity})

    raw_pattern = schema_entry.get("excel_file_pattern", "*.xlsx")
    pattern = expand_file_pattern(raw_pattern, target_date)
    add_check(
        "file_pattern",
        matches_file_pattern(filename, raw_pattern, target_date),
        f"Pattern '{pattern}' vs '{filename}'",
    )

    min_kb, size_source = resolve_min_size_kb(
        normalize_part_key(filename),
        next((s["name"] for s in schema_entry.get("sheets", [])), "Sheet1"),
        scraper_name,
        stats,
        peer_sizes_kb,
    )
    size_kb = file_size_bytes / 1024
    add_check(
        "min_file_size",
        size_kb >= min_kb,
        f"Size {size_kb:.1f} KB (min {min_kb:.1f} KB from {size_source})",
    )

    sheet_schemas = {s["name"]: s for s in schema_entry.get("sheets", [])}
    actual_sheets = inspection.get("sheets", {})
    column_schema_summary: Dict[str, Any] = {}
    known_union = stats.get(scraper_name, {}).get("column_union", [])

    for sheet_name, sheet_schema in sheet_schemas.items():
        if sheet_name not in actual_sheets:
            add_check(
                "sheet_exists",
                False,
                f"Missing sheet '{sheet_name}' (found: {list(actual_sheets.keys())})",
            )
            continue

        add_check("sheet_exists", True, f"Sheet '{sheet_name}' present")

        actual = actual_sheets[sheet_name]
        actual_cols = actual.get("columns", [])
        column_checks, column_schema = validate_column_schema(
            actual_cols, sheet_schema, known_union
        )
        column_schema_summary[sheet_name] = column_schema
        for check in column_checks:
            add_check(
                check["check"],
                check["passed"],
                check["detail"],
                severity=check.get("severity", "critical"),
            )

        part_key = normalize_part_key(filename)
        row_min, row_max, row_source = resolve_row_range(
            part_key, sheet_name, scraper_name, stats, peer_row_counts
        )
        row_count = actual.get("row_count", 0)
        if row_max is None:
            in_range = row_count >= row_min
            range_label = f">= {row_min}"
        else:
            in_range = row_min <= row_count <= row_max
            range_label = f"[{row_min}, {row_max}]"
        add_check(
            "row_count_range",
            in_range,
            f"Row count {row_count} (expected {range_label} from {row_source})",
            severity="high" if not in_range else "critical",
        )

    return checks, all_passed, column_schema_summary


def run_quality_checks(content: bytes, sheet_name: str = "Sheet1") -> List[dict]:
    checks: List[dict] = []
    df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name)

    def add(name: str, passed: bool, detail: str) -> None:
        checks.append(
            {"check": name, "passed": bool(passed), "detail": detail, "severity": "medium"}
        )

    if "ad_id" in df.columns:
        null_pct = float(df["ad_id"].isna().mean() * 100)
        add("null_ad_id_pct", null_pct < 5, f"{null_pct:.1f}% null ad_id values")
        dupes = int(df["ad_id"].duplicated().sum())
        add("duplicate_ad_id", dupes == 0, f"{dupes} duplicate ad_id values")

    if "title" in df.columns:
        null_pct = float(df["title"].isna().mean() * 100)
        add("null_title_pct", null_pct < 10, f"{null_pct:.1f}% null title values")

    if "price" in df.columns:
        null_pct = float(df["price"].isna().mean() * 100)
        add("null_price_pct", null_pct < 20, f"{null_pct:.1f}% null price values")

    return checks


def load_existing_stats(client, bucket: str, base_path: str) -> dict:
    key = f"{base_path}/{MONITOR_STATS_KEY}"
    try:
        body = download_object(client, bucket, key)
        return yaml.safe_load(body) or {}
    except ClientError as exc:
        if exc.response["Error"]["Code"] in ("NoSuchKey", "404"):
            return {}
        raise


def merge_stats(
    existing: dict,
    scraper_name: str,
    observations: List[dict],
    schema_entry: Optional[dict] = None,
) -> dict:
    stats = dict(existing)
    scraper_stats = stats.setdefault(scraper_name, {"parts": {}, "column_union": []})
    column_union = set(scraper_stats.get("column_union", []))

    for obs in observations:
        part_key = normalize_part_key(obs["filename"])
        part_stats = scraper_stats["parts"].setdefault(part_key, {})
        size_kb = obs.get("size_bytes", 0) / 1024

        for sheet_name, sheet_data in obs.get("sheets", {}).items():
            prev = part_stats.get(sheet_name, {})
            row_count = sheet_data.get("row_count", 0)
            part_stats[sheet_name] = {
                "min_rows": min(prev.get("min_rows", row_count), row_count),
                "max_rows": max(prev.get("max_rows", row_count), row_count),
                "min_size_kb": min(prev.get("min_size_kb", size_kb), size_kb),
                "max_size_kb": max(prev.get("max_size_kb", size_kb), size_kb),
                "last_seen": obs.get("date"),
                "observations": prev.get("observations", 0) + 1,
            }
            column_union.update(sheet_data.get("columns", []))

    scraper_stats["column_union"] = sorted(column_union)
    sheet_schema = (schema_entry or {}).get("sheets", [{}])[0]
    patterns = sheet_schema.get("dynamic_column_patterns", DEFAULT_DYNAMIC_COLUMN_PATTERNS)
    required_cols = sheet_schema.get("required_columns", [])
    scraper_stats["column_schema"] = {
        "required_columns": required_cols,
        "dynamic_column_patterns": patterns,
        "observed_union": sorted(column_union),
        "observed_required_count": len([c for c in column_union if c in required_cols]),
        "observed_dynamic_count": len(
            [
                c
                for c in column_union
                if c not in required_cols
                and any(fnmatch.fnmatch(c, p) for p in patterns)
            ]
        ),
        "total_observed": len(column_union),
        "last_seen": observations[-1].get("date") if observations else None,
    }
    stats[scraper_name] = scraper_stats
    return stats


def check_cross_part_column_consistency(
    pending_files: List[dict],
    sheet_schema: dict,
    primary_sheet: str,
) -> Tuple[bool, str, Dict[str, Any]]:
    """Required columns must match across parts; dynamic inspection columns may differ."""
    required_cols = sheet_schema.get("required_columns", [])
    required_set = set(required_cols)
    patterns = sheet_schema.get("dynamic_column_patterns", DEFAULT_DYNAMIC_COLUMN_PATTERNS)

    required_sets: Dict[tuple, List[str]] = {}
    part_details: List[dict] = []

    for item in pending_files:
        cols = item["inspection"]["sheets"].get(primary_sheet, {}).get("columns", [])
        part_required = sorted(c for c in cols if c in required_set)
        part_dynamic = sorted(
            c for c in cols
            if c not in required_set and any(fnmatch.fnmatch(c, p) for p in patterns)
        )
        required_sets.setdefault(tuple(part_required), []).append(item["filename"])
        part_details.append(
            {
                "filename": item["filename"],
                "total_columns": len(cols),
                "required_count": len(part_required),
                "dynamic_count": len(part_dynamic),
            }
        )

    consistency_ok = len(required_sets) <= 1
    dynamic_counts = [p["dynamic_count"] for p in part_details]
    dynamic_note = (
        f"dynamic inspection columns vary by part "
        f"({min(dynamic_counts)}–{max(dynamic_counts)} cols) — expected"
        if dynamic_counts and min(dynamic_counts) != max(dynamic_counts)
        else f"each part has {dynamic_counts[0]} dynamic inspection columns"
        if dynamic_counts
        else "no dynamic columns"
    )

    if consistency_ok:
        detail = (
            f"All {len(pending_files)} parts share {len(required_cols)} required columns; "
            f"{dynamic_note}"
        )
    else:
        mismatches = [
            f"{files[0]} (missing required: "
            f"{sorted(required_set - set(c for c in required_sets.keys() for c in c))})"
            for _, files in required_sets.items()
        ]
        detail = f"Required columns differ across parts: {'; '.join(mismatches)}"

    return consistency_ok, detail, {
        "required_columns_match": consistency_ok,
        "parts": part_details,
        "required_column_count": len(required_cols),
    }


def upload_bytes(client, bucket: str, key: str, content: bytes, content_type: str) -> None:
    client.put_object(Bucket=bucket, Key=key, Body=content, ContentType=content_type)


def collect_failures(results: dict) -> List[dict]:
    """Flatten failed checks across all scrapers/files."""
    failures: List[dict] = []
    for scraper in results.get("scrapers", []):
        consistency = scraper.get("column_consistency")
        if consistency and not consistency.get("passed"):
            failures.append(
                {
                    "scraper": scraper["name"],
                    "filename": "(column consistency)",
                    "key": "",
                    "check": "column_consistency",
                    "severity": "high",
                    "detail": consistency.get("detail", ""),
                }
            )
        for file_result in scraper.get("files", []):
            for check in file_result.get("checks", []):
                if check.get("passed"):
                    continue
                failures.append(
                    {
                        "scraper": scraper["name"],
                        "filename": file_result.get("filename", ""),
                        "key": file_result.get("key", ""),
                        "check": check.get("check", ""),
                        "severity": check.get("severity", "unknown"),
                        "detail": check.get("detail", ""),
                    }
                )
    return failures


def print_column_schema_report(scraper_name: str, scraper_result: dict) -> None:
    summary = scraper_result.get("column_schema_summary")
    if not summary:
        return

    print(f"\nColumn schema — {scraper_name}:")
    for sheet_name, sheet_summary in summary.items():
        print(f"  Sheet: {sheet_name}")
        print(f"    Total columns: {sheet_summary.get('total_columns', 0)}")
        print(
            f"    Required: {sheet_summary.get('required_count', 0)} | "
            f"Dynamic: {sheet_summary.get('dynamic_count', 0)} | "
            f"Patterns: {', '.join(sheet_summary.get('dynamic_column_patterns', []))}"
        )
        if sheet_summary.get("dynamic_columns"):
            dynamic_preview = sheet_summary["dynamic_columns"][:5]
            suffix = " ..." if len(sheet_summary["dynamic_columns"]) > 5 else ""
            print(f"    Dynamic columns: {', '.join(dynamic_preview)}{suffix}")
        if sheet_summary.get("unknown_columns"):
            print(f"    Unexpected: {sheet_summary['unknown_columns']}")
        if sheet_summary.get("new_columns"):
            print(f"    New vs history: {sheet_summary['new_columns']}")
        if sheet_summary.get("dropped_from_union"):
            print(f"    Dropped vs history: {sheet_summary['dropped_from_union']}")

    consistency = scraper_result.get("column_consistency")
    if consistency:
        status = "PASS" if consistency.get("passed") else "FAIL"
        print(f"  Cross-part required columns: {status} — {consistency.get('detail')}")


def print_file_inventory(scraper_name: str, pending_files: List[dict]) -> None:
    print(f"\nFiles found for {scraper_name}: {len(pending_files)}")
    for item in pending_files:
        sheets = item["inspection"].get("sheets", {})
        sheet_summary = ", ".join(
            f"{name}={data.get('row_count', 0)} rows"
            for name, data in sheets.items()
        )
        size_kb = item["size_bytes"] / 1024
        print(f"  - {item['filename']}")
        print(f"    key: {item['key']}")
        print(f"    size: {size_kb:.1f} KB | {sheet_summary or 'no sheets'}")


def print_failure_report(results: dict) -> None:
    failures = collect_failures(results)
    any_scraper_failed = any(not s.get("all_passed") for s in results.get("scrapers", []))
    if not failures and not any_scraper_failed:
        print("\nAll checks passed.")
        return

    print(f"\nFailed checks ({len(failures)}):")
    print("-" * 72)
    for entry in failures:
        severity = entry["severity"].upper()
        print(f"[{severity}] {entry['scraper']} :: {entry['filename']}")
        print(f"  check: {entry['check']}")
        print(f"  detail: {entry['detail']}")
        if entry["key"]:
            print(f"  key: {entry['key']}")
        print("-" * 72)


def print_per_file_status(results: dict) -> None:
    print("\nPer-file status:")
    for scraper in results.get("scrapers", []):
        for file_result in scraper.get("files", []):
            filename = file_result.get("filename", "")
            if filename.startswith("("):
                continue
            status = "PASS" if file_result.get("all_passed") else "FAIL"
            sheets = file_result.get("sheets", {})
            row_total = sum(s.get("row_count", 0) for s in sheets.values())
            size_kb = file_result.get("size_bytes", 0) / 1024
            print(
                f"  {status:<4} {filename} "
                f"({row_total} rows, {size_kb:.1f} KB)"
            )
            if not file_result.get("all_passed"):
                for check in file_result.get("checks", []):
                    if not check.get("passed"):
                        print(
                            f"       └─ [{check.get('severity', '?').upper()}] "
                            f"{check.get('check')}: {check.get('detail')}"
                        )


def write_step_summary(results: dict) -> None:
    summary_path = os.getenv("GITHUB_STEP_SUMMARY")
    if not summary_path:
        return

    lines = [
        "## R2 Excel Schema Monitor",
        "",
        "| Scraper | Files | R2 files | Unique ads | Passed | Total checks | Status |",
        "|---------|-------|----------|------------|--------|--------------|--------|",
    ]
    for scraper in results.get("scrapers", []):
        status = "PASS" if scraper.get("all_passed") else "FAIL"
        lines.append(
            f"| {scraper['name']} | {scraper['files_found']} | "
            f"{scraper.get('r2_file_count', '—')} | "
            f"{scraper.get('unique_ads', 0)} | "
            f"{scraper['checks_passed']} | {scraper['checks_total']} | {status} |"
        )
    total_unique = results.get("total_unique_ads")
    total_r2 = results.get("total_r2_files")
    if total_unique is not None:
        lines.extend(["", f"**Total unique ads:** {total_unique}", ""])
    if total_r2 is not None:
        lines.extend(["", f"**Total R2 files:** {total_r2:,}", ""])

    for scraper in results.get("scrapers", []):
        if scraper.get("all_passed"):
            continue
        lines.extend(["", f"### Failures — {scraper['name']}", ""])
        for file_result in scraper.get("files", []):
            failed = [c for c in file_result.get("checks", []) if not c["passed"]]
            if not failed:
                continue
            lines.append(f"**{file_result['filename']}**")
            if file_result.get("key"):
                lines.append(f"`{file_result['key']}`")
            for check in failed:
                lines.append(
                    f"- **[{check.get('severity', '?').upper()}]** "
                    f"`{check['check']}`: {check['detail']}"
                )
            lines.append("")

    for scraper in results.get("scrapers", []):
        summary = scraper.get("column_schema_summary")
        if not summary:
            continue
        lines.extend(["", f"### Column schema — {scraper['name']}", ""])
        for sheet_name, sheet_summary in summary.items():
            lines.append(
                f"**{sheet_name}**: {sheet_summary.get('total_columns', 0)} columns "
                f"({sheet_summary.get('required_count', 0)} required, "
                f"{sheet_summary.get('dynamic_count', 0)} dynamic)"
            )
            if sheet_summary.get("dynamic_columns"):
                preview = ", ".join(sheet_summary["dynamic_columns"][:8])
                if len(sheet_summary["dynamic_columns"]) > 8:
                    preview += ", …"
                lines.append(f"- Dynamic: `{preview}`")
        consistency = scraper.get("column_consistency")
        if consistency:
            lines.append(f"- Cross-part consistency: {consistency.get('detail')}")

    failures = collect_failures(results)
    if failures:
        lines.extend(["", "### Failure summary", ""])
        for entry in failures:
            lines.append(
                f"- **[{entry['severity'].upper()}]** `{entry['check']}` — "
                f"{entry['filename']}: {entry['detail']}"
            )

    with open(summary_path, "a", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def print_summary_table(results: dict) -> None:
    print("\nR2 Excel Schema Monitor Summary")
    print("-" * 72)
    print(
        f"{'Scraper':<24} {'Files':>6} {'R2':>8} {'Ads':>8} {'Passed':>8} {'Total':>8} {'Status':>8}"
    )
    print("-" * 72)
    for scraper in results.get("scrapers", []):
        status = "PASS" if scraper.get("all_passed") else "FAIL"
        print(
            f"{scraper['name']:<24} {scraper['files_found']:>6} "
            f"{scraper.get('r2_file_count', 0):>8} "
            f"{scraper.get('unique_ads', 0):>8} "
            f"{scraper['checks_passed']:>8} {scraper['checks_total']:>8} {status:>8}"
        )
    print("-" * 72)
    if results.get("total_unique_ads") is not None:
        print(f"Total unique ads: {results['total_unique_ads']}")
    if results.get("total_r2_files") is not None:
        print(f"Total R2 files: {results['total_r2_files']:,}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate Motorgy Excel files in R2")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    parser.add_argument("--date", default=today, help="Target date YYYY-MM-DD (UTC, default today)")
    parser.add_argument("--days-lookback", type=int, default=1, help="Number of days to inspect")
    parser.add_argument("--update-stats", action="store_true", help="Merge observations into monitor_stats.yml in R2")
    parser.add_argument("--quality", action="store_true", help="Run pandas data-quality checks")
    parser.add_argument("--fail-on-error", action="store_true", help="Exit 1 if any check fails")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    run_started_at = datetime.now(timezone.utc)
    bucket = get_env("CF_R2_BUCKET_NAME")
    client = build_r2_client()
    config_base = os.getenv("CONFIG_R2_BASE", DEFAULT_CONFIG_BASE)
    config = load_config_from_r2(client, bucket, config_base)

    schema_cfg = {s["scraper"]: s for s in config.get("excel_schema", [])}

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "target_dates": [],
        "scrapers": [],
    }
    report_base = strip_bucket_placeholder(config.get("scrapers", [{}])[0].get("r2_path", "motorgy"))
    stats_merged = load_existing_stats(client, bucket, report_base)
    any_failure = False

    target_date = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    dates = [target_date - timedelta(days=offset) for offset in range(args.days_lookback)]
    report["target_dates"] = [d.strftime("%Y-%m-%d") for d in dates]
    print(f"Validation target date(s): {', '.join(report['target_dates'])}")
    print(f"Quality checks: {'enabled' if args.quality else 'disabled'}")

    for scraper in config.get("scrapers", []):
        scraper_name = scraper["name"]
        schema_entry = schema_cfg.get(scraper_name)
        if not schema_entry:
            print(f"WARNING: No excel_schema entry for scraper '{scraper_name}'", file=sys.stderr)
            continue

        base_path = strip_bucket_placeholder(scraper["r2_path"])
        scraper_result = {
            "name": scraper_name,
            "files_found": 0,
            "checks_passed": 0,
            "checks_total": 0,
            "all_passed": True,
            "files": [],
        }
        stat_observations: List[dict] = []
        pending_files: List[dict] = []

        for day in dates:
            prefix = date_partition_prefix(base_path, day)
            print(f"Listing R2 prefix: {prefix}/")
            objects = list_xlsx_objects(client, bucket, prefix)
            print(f"  Found {len(objects)} .xlsx object(s)")

            for obj in objects:
                key = obj["Key"]
                filename = os.path.basename(key)
                if "/excel_files/" not in key:
                    continue

                content = download_object(client, bucket, key)
                inspection = inspect_workbook(content)
                pending_files.append(
                    {
                        "key": key,
                        "filename": filename,
                        "size_bytes": obj["Size"],
                        "inspection": inspection,
                        "content": content,
                        "day": day,
                    }
                )

        print_file_inventory(scraper_name, pending_files)

        scraper_result["files_found"] = len(pending_files)
        known_parts = set(stats_merged.get(scraper_name, {}).get("parts", {}).keys())
        current_parts = {normalize_part_key(item["filename"]) for item in pending_files}
        if known_parts:
            missing_parts = sorted(known_parts - current_parts)
            parts_ok = len(missing_parts) == 0
            scraper_result["checks_total"] += 1
            scraper_result["checks_passed"] += int(parts_ok)
            if not parts_ok:
                scraper_result["all_passed"] = False
                any_failure = True
                print(
                    f"  FAIL (parts completeness): missing {missing_parts} "
                    f"(expected {sorted(known_parts)})"
                )
            else:
                print(f"  PASS (parts completeness): all {len(known_parts)} known parts present")
            scraper_result["files"].append(
                {
                    "key": "",
                    "filename": "(parts completeness)",
                    "checks": [
                        {
                            "check": "parts_completeness",
                            "passed": parts_ok,
                            "detail": (
                                f"Missing parts: {missing_parts} "
                                f"(expected {sorted(known_parts)} from R2 history)"
                                if missing_parts
                                else f"All {len(known_parts)} known parts present"
                            ),
                            "severity": "critical",
                        }
                    ],
                    "all_passed": parts_ok,
                }
            )

        peer_row_counts = [
            sheet.get("row_count", 0)
            for item in pending_files
            for sheet in item["inspection"].get("sheets", {}).values()
        ]
        peer_sizes_kb = [item["size_bytes"] / 1024 for item in pending_files]
        known_union = stats_merged.get(scraper_name, {}).get("column_union", [])
        sheet_schema = schema_entry.get("sheets", [{}])[0]
        primary_sheet = sheet_schema.get("name", "Sheet1")

        for item in pending_files:
            checks, file_passed, file_column_schema = validate_file(
                item["filename"],
                item["size_bytes"],
                item["inspection"],
                schema_entry,
                scraper_name,
                stats_merged,
                peer_row_counts,
                peer_sizes_kb,
                item["day"],
            )

            if args.quality:
                quality_checks = run_quality_checks(item["content"])
                checks.extend(quality_checks)
                file_passed = file_passed and all(c["passed"] for c in quality_checks)

            scraper_result["checks_total"] += len(checks)
            scraper_result["checks_passed"] += sum(1 for c in checks if c["passed"])
            if not file_passed:
                scraper_result["all_passed"] = False
                any_failure = True
                failed_names = [c["check"] for c in checks if not c["passed"]]
                print(f"  FAIL {item['filename']}: {', '.join(failed_names)}")
            else:
                print(f"  PASS {item['filename']}")

            scraper_result["files"].append(
                {
                    "key": item["key"],
                    "filename": item["filename"],
                    "size_bytes": int(item["size_bytes"]),
                    "sheets": item["inspection"].get("sheets", {}),
                    "column_schema": file_column_schema,
                    "checks": checks,
                    "all_passed": bool(file_passed),
                }
            )
            stat_observations.append(
                {
                    "filename": item["filename"],
                    "date": item["day"].strftime("%Y-%m-%d"),
                    "size_bytes": item["size_bytes"],
                    "sheets": item["inspection"].get("sheets", {}),
                }
            )

        if pending_files:
            all_columns: List[str] = []
            for item in pending_files:
                cols = item["inspection"]["sheets"].get(primary_sheet, {}).get("columns", [])
                all_columns.extend(col for col in cols if col not in all_columns)

            scraper_result["column_schema_summary"] = {
                primary_sheet: classify_columns(all_columns, sheet_schema, known_union)
            }

            consistency_ok, consistency_detail, consistency_meta = check_cross_part_column_consistency(
                pending_files, sheet_schema, primary_sheet
            )
            scraper_result["column_consistency"] = {
                "passed": consistency_ok,
                "detail": consistency_detail,
                **consistency_meta,
            }
            if len(pending_files) > 1:
                scraper_result["checks_total"] += 1
                scraper_result["checks_passed"] += int(consistency_ok)
                if not consistency_ok:
                    scraper_result["all_passed"] = False
                    any_failure = True
                    print(f"  FAIL (column consistency): {consistency_detail}")
                else:
                    print(f"  PASS (column consistency): {consistency_detail}")

            print_column_schema_report(scraper_name, scraper_result)

        excel_downloads = [(item["key"], item["content"]) for item in pending_files]
        ads_stats = count_scraper_ads(
            client, bucket, base_path, target_date, excel_downloads
        )
        scraper_result["unique_ads"] = ads_stats.get("unique_ads") or 0
        scraper_result["total_rows"] = ads_stats.get("total_rows") or 0
        scraper_result["ads_source"] = ads_stats.get("ads_source", "none")
        print(
            f"  Unique ads: {scraper_result['unique_ads']} "
            f"(source: {scraper_result['ads_source']})"
        )

        scraper_result["r2_file_count"] = count_scraper_r2_files(
            client, bucket, base_path
        )

        req_stats = count_scraper_request_metrics(
            client, bucket, base_path, target_date
        )
        scraper_result["requests_total"] = req_stats.get("requests_total")
        scraper_result["requests_failed"] = req_stats.get("requests_failed")
        scraper_result["error_rate_pct"] = req_stats.get("error_rate_pct")
        scraper_result["requests_per_min"] = req_stats.get("requests_per_min")
        scraper_result["duration_sec"] = req_stats.get("duration_sec")
        scraper_result["metrics_source"] = req_stats.get("metrics_source", "none")
        if req_stats.get("failed_items_summary"):
            scraper_result["failed_items_summary"] = req_stats["failed_items_summary"]
        if req_stats.get("metrics_source") != "none":
            print(
                f"  Request metrics: {scraper_result.get('requests_total', 0)} total, "
                f"{scraper_result.get('requests_failed', 0)} failed, "
                f"{scraper_result.get('error_rate_pct', 0)}% error rate"
            )

        if scraper_result["files_found"] == 0:
            scraper_result["all_passed"] = False
            any_failure = True
            scraper_result["checks_total"] += 1
            scraper_result["files"].append(
                {
                    "key": "",
                    "filename": "(none)",
                    "checks": [
                        {
                            "check": "files_found",
                            "passed": False,
                            "detail": f"No .xlsx files under {base_path} for target dates",
                            "severity": "critical",
                        }
                    ],
                    "all_passed": False,
                }
            )

        report["scrapers"].append(scraper_result)

        if stat_observations:
            stats_merged = merge_stats(
                stats_merged, scraper_name, stat_observations, schema_entry
            )

    report["total_unique_ads"] = sum(
        r.get("unique_ads") or 0 for r in report["scrapers"]
    )

    site_r2_prefix = (config.get("r2_prefix") or report_base).strip("/")
    if site_r2_prefix:
        report["total_r2_files"] = count_site_r2_files(
            client, bucket, site_r2_prefix
        )
    else:
        report["total_r2_files"] = sum(
            r.get("r2_file_count") or 0 for r in report["scrapers"]
        )

    site_metrics = aggregate_site_request_metrics(report["scrapers"])
    report.update(site_metrics)
    report["error_summary"] = build_run_error_summary(report["scrapers"], [])

    site_meta = load_site_run_meta()
    report["github_run"] = build_scraper_run_meta(
        site_meta,
        args.date,
        run_started_at.replace(tzinfo=None),
        not any_failure,
    )
    report["run_place"] = report["github_run"].get("run_place")

    report_key = f"{report_base}/monitor/{args.date}/report.json"
    upload_bytes(
        client,
        bucket,
        report_key,
        json.dumps(json_safe(report), indent=2, ensure_ascii=False).encode("utf-8"),
        "application/json",
    )
    print(f"Report uploaded to r2://{bucket}/{report_key}")

    if stats_merged:
        stats_key = f"{report_base}/{MONITOR_STATS_KEY}"
        stats_body = yaml.safe_dump(stats_merged, sort_keys=False, allow_unicode=True).encode("utf-8")
        upload_bytes(client, bucket, stats_key, stats_body, "text/yaml")
        print(f"Stats uploaded to r2://{bucket}/{stats_key}")

    print_summary_table(report)
    print_per_file_status(report)
    print_failure_report(report)
    write_step_summary(report)

    if any_failure:
        print("\nValidation result: FAIL")
    else:
        print("\nValidation result: PASS")

    if args.fail_on_error and any_failure:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
